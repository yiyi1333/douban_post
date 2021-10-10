# coding=utf-8
import random

import requests
from xlutils.copy import copy
import xlwt
import time
import openpyxl
import xlrd
from bs4 import BeautifulSoup

def getimgpath(movid, num):
    url = 'https://movie.douban.com/subject/' + movid + '/'
    proxieslist = [
        {
            'http': 'http://106.110.156.183:4247'
        },
        {
            'http': 'http://60.184.97.235:4245'
        },

        {
            'http': 'http://114.96.218.79:4281'
        },
        {
            'http': 'http://36.6.68.102:4245'
        },
        {
            'http': 'http://115.207.17.37:4245'
        },
        {
            'http': 'http://114.239.106.70:4245'
        },
        {
            'http': 'http://183.166.7.38:4231'
        },
        {
            'http': 'http://115.198.59.150:4286'
        },
        {
            'http': 'http://61.188.26.16:4210'
        },
        {
            'http': 'http://117.70.35.33:4210'
        }
    ]

    # proxy = random.choice(proxies)
    headerlist = [
        {'User-agent':'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'},
        {'User-agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.153 Safari/537.36'},
        {'User-agent':'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:30.0) Gecko/20100101 Firefox/30.0'},
        {'User-agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/537.75.14'},
        {'User-agent':'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Win64; x64; Trident/6.0)'},
        {'User-agent':'Mozilla/5.0 (Windows; U; Windows NT 5.1; it; rv:1.8.1.11) Gecko/20071127 Firefox/2.0.0.11'},
        {'User-agent':'Opera/9.25 (Windows NT 5.1; U; en)'},
        {'User-agent':'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)'},
        {'User-agent':'Mozilla/5.0 (compatible; Konqueror/3.5; Linux) KHTML/3.5.5 (like Gecko) (Kubuntu)'},
        {'User-agent':'Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.8.0.12) Gecko/20070731 Ubuntu/dapper-security Firefox/1.5.0.12'},
        {'User-agent':'Lynx/2.8.5rel.1 libwww-FM/2.14 SSL-MM/1.4.1 GNUTLS/1.2.9'},
        {'User-agent':'Mozilla/5.0 (X11; Linux i686) AppleWebKit/535.7 (KHTML, like Gecko) Ubuntu/11.04 Chromium/16.0.912.77 Chrome/16.0.912.77 Safari/535.7'},
        {'User-agent':'Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:10.0) Gecko/20100101 Firefox/10.0 '}
    ]
    headers = random.choice(headerlist)
    proxies = random.choice(proxieslist)
    res = requests.get(url, headers=headers)

    if(res.status_code == 200):
        ans = ''
        try:
            soup = BeautifulSoup(res.text, 'lxml')
            ans = soup.find(attrs={'id': 'mainpic'}).find('img').get('src')
        except BaseException:
            print('Error: ', num)
            return 'Error'
        else:
            print('(', num, '/', 28603, ')', '[ url:', ans, ']')
            return ans
    else:
        print('状态码：',res.status_code)
        return 'none'

def write_excel(file, line, id, url):
    #文件读取
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.worksheets[0]
    #修改数据
    for i in range(len(id)):
        worksheet.cell(line + i, 1, id[i])
        worksheet.cell(line + i, 2, url[i])
    #保存
    workbook.save(filename=file)

def read_excel(file, begin, n):
    wb = xlrd.open_workbook(filename=file)
    sheet1 = wb.sheet_by_index(0)
    rows = []
    for i in range(n):
        row = sheet1.row_values(begin + i)
        rows.append(row[0])
    return rows


beginline = 10496
n = 10
failnum = 0
while (beginline < 20000):
    # 读取n条id到idlist
    idlist = read_excel('testdb_MovieInfo.xlsx', beginline, n)
    # 爬取n条url到imglist
    imglist = []
    for i in range(n):
        img = getimgpath(idlist[i], beginline + i)
        if(img == 'none'):
            failnum += 1
            print('失败数：', failnum)
        imglist.append(img)
        time.sleep(2)
    write_excel('movurl.xlsx', beginline + 1, idlist, imglist)
    beginline += n
    print('Nowline_input: ',beginline)

